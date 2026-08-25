from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.optimization_engine import generate_optimization_candidates


def format_header(worksheet):

    for cell in worksheet[1]:

        cell.font = Font(bold=True)

        cell.alignment = Alignment(
            horizontal="center"
        )


def auto_adjust_columns(worksheet):

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            50
        )


def add_table(
    worksheet,
    table_name
):

    if worksheet.max_row < 2:

        return

    reference = (
        f"A1:"
        f"{get_column_letter(worksheet.max_column)}"
        f"{worksheet.max_row}"
    )

    table = Table(
        displayName=table_name,
        ref=reference
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style

    worksheet.add_table(table)


def generate_excel_report(
    output_path,
    df,
    kpi_summary,
    validation_summary,
    reconciliation_result,
    cpu_threshold,
    savings_assumption,
    anomaly_summary=None,
    anomalies=None,
    savings_intelligence=None,
    action_register=None
):

    output_file = Path(output_path)

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook = Workbook()

    # ========================================
    # 1. EXECUTIVE SUMMARY
    # ========================================

    summary_sheet = workbook.active

    summary_sheet.title = "Executive Summary"

    summary_sheet.append([
        "FinOps KPI",
        "Value"
    ])

    summary_sheet.append([
        "Total Cloud Spend",
        kpi_summary["total_spend"]
    ])

    summary_sheet.append([
        "Resource Count",
        kpi_summary["resource_count"]
    ])

    summary_sheet.append([
        "Average Resource Cost",
        kpi_summary["average_resource_cost"]
    ])

    summary_sheet.append([
        "Running Resources",
        kpi_summary["running_resources"]
    ])

    summary_sheet.append([
        "Running Cost",
        kpi_summary["running_cost"]
    ])

    summary_sheet.append([
        "Stopped Resources",
        kpi_summary["stopped_resources"]
    ])

    summary_sheet.append([
        "Stopped Cost Exposure",
        kpi_summary["stopped_cost"]
    ])

    summary_sheet.append([
        "Service Count",
        kpi_summary["service_count"]
    ])

    summary_sheet.append([
        "Business Unit Count",
        kpi_summary["business_unit_count"]
    ])

    summary_sheet.append([
        "Average CPU Utilization",
        kpi_summary["average_cpu_utilization"]
    ])

    summary_sheet.append([
        "Total Optimization Investigations",
        kpi_summary["optimization_candidates"]
    ])

    summary_sheet.append([
        "Rightsizing Candidates",
        kpi_summary["rightsizing_candidates"]
    ])

    summary_sheet.append([
        "Rightsizing Cost Exposure",
        kpi_summary[
            "rightsizing_cost_exposure"
        ]
    ])

    summary_sheet.append([
        "Estimated Rightsizing Savings",
        kpi_summary[
            "estimated_rightsizing_savings"
        ]
    ])

    summary_sheet.append([
        "Estimated Savings %",
        kpi_summary[
            "estimated_savings_percentage"
        ] / 100
    ])

    summary_sheet.append([
        "Stopped Resource Investigations",
        kpi_summary[
            "stopped_resource_investigations"
        ]
    ])

    summary_sheet.append([
        "Stopped Cost Exposure",
        kpi_summary[
            "stopped_cost_exposure"
        ]
    ])

    summary_sheet.append([
        "High Priority Opportunities",
        kpi_summary[
            "high_priority_opportunities"
        ]
    ])

    summary_sheet.append([
        "Medium Priority Opportunities",
        kpi_summary[
            "medium_priority_opportunities"
        ]
    ])

    format_header(
        summary_sheet
    )

    summary_sheet.freeze_panes = "A2"

    # Currency

    currency_rows = [
        2,
        4,
        6,
        8,
        14,
        15,
        18
    ]

    for row in currency_rows:

        summary_sheet[
            f"B{row}"
        ].number_format = '₹#,##0.00'

    # CPU utilization

    summary_sheet[
        "B11"
    ].number_format = "0.00"

    # Savings percentage

    summary_sheet[
        "B16"
    ].number_format = "0.00%"

    auto_adjust_columns(
        summary_sheet
    )

    add_table(
        summary_sheet,
        "ExecutiveKPI"
    )

    # ========================================
    # 2. SERVICE COST
    # ========================================

    service_sheet = workbook.create_sheet(
        "Service Cost"
    )

    service_sheet.append([
        "Service",
        "Monthly Cost",
        "% of Total Spend"
    ])

    service_cost = (
        df.groupby("Service")["Monthly_Cost"]
        .sum()
        .sort_values(ascending=False)
    )

    total_spend = float(
        df["Monthly_Cost"].sum()
    )

    for service, cost in service_cost.items():

        cost = float(cost)

        service_sheet.append([
            service,
            cost,
            cost / total_spend
        ])

    format_header(
        service_sheet
    )

    service_sheet.freeze_panes = "A2"

    for row in range(
        2,
        service_sheet.max_row + 1
    ):

        service_sheet[
            f"B{row}"
        ].number_format = '₹#,##0.00'

        service_sheet[
            f"C{row}"
        ].number_format = '0.00%'

    auto_adjust_columns(
        service_sheet
    )

    add_table(
        service_sheet,
        "ServiceCost"
    )

    # ========================================
    # 3. BUSINESS UNIT COST
    # ========================================

    business_sheet = workbook.create_sheet(
        "Business Unit Cost"
    )

    business_sheet.append([
        "Business Unit",
        "Monthly Cost",
        "% of Total Spend"
    ])

    business_cost = (
        df.groupby("Business_Unit")[
            "Monthly_Cost"
        ]
        .sum()
        .sort_values(ascending=False)
    )

    for business_unit, cost in business_cost.items():

        cost = float(cost)

        business_sheet.append([
            business_unit,
            cost,
            cost / total_spend
        ])

    format_header(
        business_sheet
    )

    business_sheet.freeze_panes = "A2"

    for row in range(
        2,
        business_sheet.max_row + 1
    ):

        business_sheet[
            f"B{row}"
        ].number_format = '₹#,##0.00'

        business_sheet[
            f"C{row}"
        ].number_format = '0.00%'

    auto_adjust_columns(
        business_sheet
    )

    add_table(
        business_sheet,
        "BusinessUnitCost"
    )

    # ========================================
    # 4. OPTIMIZATION OPPORTUNITIES
    # ========================================

    optimization_sheet = workbook.create_sheet(
        "Optimization Opportunities"
    )

    optimization_sheet.append([
        "Rank",
        "Resource ID",
        "Service",
        "Resource Status",
        "Business Unit",
        "Environment",
        "CPU Utilization",
        "Monthly Cost",
        "Estimated Savings",
        "Opportunity Type",
        "Reason",
        "Recommendation",
        "Priority"
    ])

    optimization_candidates = (
        generate_optimization_candidates(
            df,
            cpu_threshold,
            savings_assumption
        )
    )

    for _, row in (
        optimization_candidates.iterrows()
    ):

        optimization_sheet.append([
            row["Optimization_Rank"],
            row["Resource_ID"],
            row["Service"],
            row["Resource_Status"],
            row["Business_Unit"],
            row["Environment"],
            float(row["CPU_Utilization"]),
            float(row["Monthly_Cost"]),
            float(row["Estimated_Savings"]),
            row["Opportunity_Type"],
            row["Reason"],
            row["Recommendation"],
            row["Priority"]
        ])

    format_header(
        optimization_sheet
    )

    optimization_sheet.freeze_panes = "A2"

    for row in range(
        2,
        optimization_sheet.max_row + 1
    ):

        optimization_sheet[
            f"G{row}"
        ].number_format = "0.00"

        optimization_sheet[
            f"H{row}"
        ].number_format = '₹#,##0.00'

        optimization_sheet[
            f"I{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        optimization_sheet
    )

    add_table(
        optimization_sheet,
        "OptimizationCandidates"
    )

    # ========================================
    # 5. DATA QUALITY
    # ========================================

    quality_sheet = workbook.create_sheet(
        "Data Quality"
    )

    quality_sheet.append([
        "Validation Check",
        "Status"
    ])

    for check, status in (
        validation_summary.items()
    ):

        quality_sheet.append([
            check,
            status
        ])

    format_header(
        quality_sheet
    )

    quality_sheet.freeze_panes = "A2"

    auto_adjust_columns(
        quality_sheet
    )

    add_table(
        quality_sheet,
        "DataQuality"
    )

    # ========================================
    # 6. RECONCILIATION
    # ========================================

    reconciliation_sheet = workbook.create_sheet(
        "Reconciliation"
    )

    reconciliation_sheet.append([
        "Metric",
        "Value"
    ])

    reconciliation_sheet.append([
        "Expected Total",
        reconciliation_result[
            "expected_total"
        ]
    ])

    reconciliation_sheet.append([
        "Processed Total",
        reconciliation_result[
            "processed_total"
        ]
    ])

    reconciliation_sheet.append([
        "Difference",
        reconciliation_result[
            "difference"
        ]
    ])

    reconciliation_sheet.append([
        "Tolerance",
        reconciliation_result[
            "tolerance"
        ]
    ])

    reconciliation_sheet.append([
        "Status",
        reconciliation_result[
            "status"
        ]
    ])

    format_header(
        reconciliation_sheet
    )

    reconciliation_sheet.freeze_panes = "A2"

    for row in range(2, 6):

        reconciliation_sheet[
            f"B{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        reconciliation_sheet
    )

    add_table(
        reconciliation_sheet,
        "Reconciliation"
    )



    # ========================================
    # 7. ANOMALY SUMMARY
    # ========================================

    anomaly_summary_sheet = workbook.create_sheet(
        "Anomaly Summary"
    )

    anomaly_summary_sheet.append([
        "Anomaly Metric",
        "Value"
    ])

    if anomaly_summary is None:
        anomaly_summary = {
            "total_anomalies": 0,
            "critical_anomalies": 0,
            "high_anomalies": 0,
            "high_cost_anomalies": 0,
            "low_utilization_high_cost": 0,
            "stopped_high_cost": 0
        }

    anomaly_summary_sheet.append([
        "Total Anomalies",
        anomaly_summary["total_anomalies"]
    ])

    anomaly_summary_sheet.append([
        "Critical Anomalies",
        anomaly_summary["critical_anomalies"]
    ])

    anomaly_summary_sheet.append([
        "High Anomalies",
        anomaly_summary["high_anomalies"]
    ])

    anomaly_summary_sheet.append([
        "High Cost Anomalies",
        anomaly_summary["high_cost_anomalies"]
    ])

    anomaly_summary_sheet.append([
        "Low Utilization + High Cost",
        anomaly_summary["low_utilization_high_cost"]
    ])

    anomaly_summary_sheet.append([
        "Stopped Resource + High Cost",
        anomaly_summary["stopped_high_cost"]
    ])

    format_header(
        anomaly_summary_sheet
    )

    anomaly_summary_sheet.freeze_panes = "A2"

    auto_adjust_columns(
        anomaly_summary_sheet
    )

    add_table(
        anomaly_summary_sheet,
        "AnomalySummary"
    )


    # ========================================
    # 8. ANOMALY DETAILS
    # ========================================

    anomaly_details_sheet = workbook.create_sheet(
        "Anomaly Details"
    )

    anomaly_details_sheet.append([
        "Rank",
        "Resource ID",
        "Service",
        "Business Unit",
        "Environment",
        "Resource Status",
        "CPU Utilization",
        "Monthly Cost",
        "Anomaly Type",
        "Severity",
        "Observed Value",
        "Threshold",
        "Recommendation"
    ])

    if anomalies is not None and not anomalies.empty:

        for _, row in anomalies.iterrows():

            anomaly_details_sheet.append([
                int(row["Anomaly_Rank"]),
                row["Resource_ID"],
                row["Service"],
                row["Business_Unit"],
                row.get("Environment", ""),
                row["Resource_Status"],
                float(row["CPU_Utilization"]),
                float(row["Monthly_Cost"]),
                str(row["Anomaly_Type"]),
                str(row["Severity"]),
                float(row["Observed_Value"]),
                float(row["Threshold"]),
                str(row["Recommendation"])
            ])

    format_header(
        anomaly_details_sheet
    )

    anomaly_details_sheet.freeze_panes = "A2"

    for row in range(
        2,
        anomaly_details_sheet.max_row + 1
    ):

        anomaly_details_sheet[
            f"G{row}"
        ].number_format = "0.00"

        anomaly_details_sheet[
            f"H{row}"
        ].number_format = '₹#,##0.00'

        anomaly_details_sheet[
            f"K{row}"
        ].number_format = '₹#,##0.00'

        anomaly_details_sheet[
            f"L{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        anomaly_details_sheet
    )

    add_table(
        anomaly_details_sheet,
        "AnomalyDetails"
    )


    # ========================================
    # 9. SAVINGS INTELLIGENCE
    # ========================================

    savings_intelligence_sheet = workbook.create_sheet(
        "Savings Intelligence"
    )

    savings_intelligence_sheet.append([
        "Savings Metric",
        "Value"
    ])

    if savings_intelligence is not None:

        savings_intelligence_sheet.append([
            "Total Spend",
            float(
                savings_intelligence[
                    "total_spend"
                ]
            )
        ])

        savings_intelligence_sheet.append([
            "Optimization Exposure",
            float(
                savings_intelligence[
                    "optimization_exposure"
                ]
            )
        ])

        savings_intelligence_sheet.append([
            "Rightsizing Exposure",
            float(
                savings_intelligence[
                    "rightsizing_exposure"
                ]
            )
        ])

        savings_intelligence_sheet.append([
            "Estimated Savings",
            float(
                savings_intelligence[
                    "estimated_savings"
                ]
            )
        ])

        savings_intelligence_sheet.append([
            "Savings Percentage",
            float(
                savings_intelligence[
                    "savings_percentage"
                ]
            )
        ])

        savings_intelligence_sheet.append([
            "Rightsizing Candidates",
            int(
                savings_intelligence[
                    "rightsizing_count"
                ]
            )
        ])

    format_header(
        savings_intelligence_sheet
    )

    savings_intelligence_sheet.freeze_panes = "A2"

    for row in range(
        2,
        savings_intelligence_sheet.max_row + 1
    ):

        metric = savings_intelligence_sheet[
            f"A{row}"
        ].value

        if metric in [
            "Total Spend",
            "Optimization Exposure",
            "Rightsizing Exposure",
            "Estimated Savings"
        ]:

            savings_intelligence_sheet[
                f"B{row}"
            ].number_format = '₹#,##0.00'

        elif metric == "Savings Percentage":

            savings_intelligence_sheet[
                f"B{row}"
            ].number_format = "0.00"

    auto_adjust_columns(
        savings_intelligence_sheet
    )

    add_table(
        savings_intelligence_sheet,
        "SavingsIntelligence"
    )


    # ========================================
    # 10. SERVICE SAVINGS
    # ========================================

    service_savings_sheet = workbook.create_sheet(
        "Service Savings"
    )

    service_savings_sheet.append([
        "Service",
        "Resource Count",
        "Cost Exposure",
        "Estimated Savings"
    ])

    if savings_intelligence is not None:

        service_savings = (
            savings_intelligence[
                "service_savings"
            ]
        )

        if (
            service_savings is not None
            and not service_savings.empty
        ):

            for service, row in service_savings.iterrows():

                service_savings_sheet.append([
                    service,
                    int(
                        row["resource_count"]
                    ),
                    float(
                        row["cost_exposure"]
                    ),
                    float(
                        row["estimated_savings"]
                    )
                ])

    format_header(
        service_savings_sheet
    )

    service_savings_sheet.freeze_panes = "A2"

    for row in range(
        2,
        service_savings_sheet.max_row + 1
    ):

        service_savings_sheet[
            f"C{row}"
        ].number_format = '₹#,##0.00'

        service_savings_sheet[
            f"D{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        service_savings_sheet
    )

    add_table(
        service_savings_sheet,
        "ServiceSavings"
    )


    # ========================================
    # 11. BUSINESS UNIT SAVINGS
    # ========================================

    business_unit_savings_sheet = workbook.create_sheet(
        "Business Unit Savings"
    )

    business_unit_savings_sheet.append([
        "Business Unit",
        "Resource Count",
        "Cost Exposure",
        "Estimated Savings"
    ])

    if savings_intelligence is not None:

        business_unit_savings = (
            savings_intelligence[
                "business_unit_savings"
            ]
        )

        if (
            business_unit_savings is not None
            and not business_unit_savings.empty
        ):

            for business_unit, row in (
                business_unit_savings.iterrows()
            ):

                business_unit_savings_sheet.append([
                    business_unit,
                    int(
                        row["resource_count"]
                    ),
                    float(
                        row["cost_exposure"]
                    ),
                    float(
                        row["estimated_savings"]
                    )
                ])

    format_header(
        business_unit_savings_sheet
    )

    business_unit_savings_sheet.freeze_panes = "A2"

    for row in range(
        2,
        business_unit_savings_sheet.max_row + 1
    ):

        business_unit_savings_sheet[
            f"C{row}"
        ].number_format = '₹#,##0.00'

        business_unit_savings_sheet[
            f"D{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        business_unit_savings_sheet
    )

    add_table(
        business_unit_savings_sheet,
        "BusinessUnitSavings"
    )


    # ========================================
    # 12. TOP SAVINGS OPPORTUNITIES
    # ========================================

    top_savings_sheet = workbook.create_sheet(
        "Top Savings Opportunities"
    )

    top_savings_sheet.append([
        "Rank",
        "Resource ID",
        "Service",
        "Business Unit",
        "Resource Status",
        "CPU Utilization",
        "Monthly Cost",
        "Estimated Savings",
        "Priority",
        "Recommendation"
    ])

    if savings_intelligence is not None:

        top_opportunities = (
            savings_intelligence[
                "top_opportunities"
            ]
        )

        if (
            top_opportunities is not None
            and not top_opportunities.empty
        ):

            for _, row in top_opportunities.iterrows():

                top_savings_sheet.append([
                    int(
                        row["Savings_Rank"]
                    ),
                    row["Resource_ID"],
                    row["Service"],
                    row["Business_Unit"],
                    row["Resource_Status"],
                    float(
                        row["CPU_Utilization"]
                    ),
                    float(
                        row["Monthly_Cost"]
                    ),
                    float(
                        row["Estimated_Savings"]
                    ),
                    row["Priority"],
                    row["Recommendation"]
                ])

    format_header(
        top_savings_sheet
    )

    top_savings_sheet.freeze_panes = "A2"

    for row in range(
        2,
        top_savings_sheet.max_row + 1
    ):

        top_savings_sheet[
            f"F{row}"
        ].number_format = "0.00"

        top_savings_sheet[
            f"G{row}"
        ].number_format = '₹#,##0.00'

        top_savings_sheet[
            f"H{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        top_savings_sheet
    )

    add_table(
        top_savings_sheet,
        "TopSavingsOpportunities"
    )

        # ========================================
    # 13. SAVINGS ACTION REGISTER
    # ========================================

    action_register_sheet = workbook.create_sheet(
        "Savings Action Register"
    )

    action_register_sheet.append([
        "Action ID",
        "Savings Rank",
        "Resource ID",
        "Service",
        "Business Unit",
        "Owner",
        "Priority",
        "Resource Status",
        "CPU Utilization",
        "Monthly Cost",
        "Estimated Savings",
        "Recommendation",
        "Action Status",
        "Target Date",
        "Notes"
    ])

    if (
        action_register is not None
        and not action_register.empty
    ):

        for _, row in action_register.iterrows():

            action_register_sheet.append([
                row["Action_ID"],
                int(
                    row["Savings_Rank"]
                ),
                row["Resource_ID"],
                row["Service"],
                row["Business_Unit"],
                row["Owner"],
                row["Priority"],
                row["Resource_Status"],
                float(
                    row["CPU_Utilization"]
                ),
                float(
                    row["Monthly_Cost"]
                ),
                float(
                    row["Estimated_Savings"]
                ),
                row["Recommendation"],
                row["Action_Status"],
                row["Target_Date"],
                row["Notes"]
            ])

    format_header(
        action_register_sheet
    )

    action_register_sheet.freeze_panes = "A2"

    for row in range(
        2,
        action_register_sheet.max_row + 1
    ):

        action_register_sheet[
            f"I{row}"
        ].number_format = "0.00"

        action_register_sheet[
            f"J{row}"
        ].number_format = '₹#,##0.00'

        action_register_sheet[
            f"K{row}"
        ].number_format = '₹#,##0.00'

    auto_adjust_columns(
        action_register_sheet
    )

    add_table(
        action_register_sheet,
        "SavingsActionRegister"
    )


    # ========================================
    # SAVE
    # ========================================

    workbook.save(
        output_file
    )

    return output_file
